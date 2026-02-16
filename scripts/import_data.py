#!/usr/bin/env python3
"""
Script to import data from Excel files into the database via API
"""
import json
import requests
import sys
from openpyxl import load_workbook

API_URL = "http://localhost:8080/api"

def import_gap_assessments(file_path):
    """Import gap assessments from Excel file"""
    print(f"Importing gap assessments from {file_path}...")
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb['Requirements']
    
    # Skip header rows (first 3 rows)
    imported = 0
    errors = 0
    
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        if not row[0]:  # Skip empty rows
            continue
        
        data = {
            'category': row[0] or '',
            'section': row[1] or '',
            'standard_ref': row[2] or '',
            'assessment_question': row[3] or '',
            'compliance': row[4] or 'Not Compliant',
            'notes': row[5] or ''
        }
        
        try:
            response = requests.post(f"{API_URL}/gap-assessments", json=data)
            if response.status_code == 201:
                imported += 1
            else:
                print(f"Error importing row: {response.status_code} - {response.text}")
                errors += 1
        except Exception as e:
            print(f"Error: {e}")
            errors += 1
    
    wb.close()
    print(f"Imported {imported} gap assessments. Errors: {errors}")

def import_maturity_assessments(file_path):
    """Import maturity assessments from Excel file"""
    print(f"Importing maturity assessments from {file_path}...")
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb['Requirements']
    
    imported = 0
    errors = 0
    
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if not row[0]:  # Skip empty rows
            continue
        
        # Handle score conversion
        current_score = None
        target_score = None
        try:
            if row[5]:
                current_score = int(row[5]) if isinstance(row[5], (int, float)) else None
        except:
            pass
        try:
            if row[8]:
                target_score = int(row[8]) if isinstance(row[8], (int, float)) else None
        except:
            pass
        
        data = {
            'category': row[0] or '',
            'section': row[1] or '',
            'standard_ref': row[2] or '',
            'assessment_question': row[3] or '',
            'current_maturity_level': str(row[4]) if row[4] else '',
            'current_maturity_score': current_score,
            'current_maturity_comments': str(row[6]) if row[6] else '',
            'target_maturity_level': str(row[7]) if row[7] else '',
            'target_maturity_score': target_score,
            'target_maturity_comments': str(row[9]) if row[9] else ''
        }
        
        try:
            response = requests.post(f"{API_URL}/maturity-assessments", json=data)
            if response.status_code == 201:
                imported += 1
            else:
                print(f"Error importing row: {response.status_code} - {response.text}")
                errors += 1
        except Exception as e:
            print(f"Error: {e}")
            errors += 1
    
    wb.close()
    print(f"Imported {imported} maturity assessments. Errors: {errors}")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python3 import_data.py <gap_file.xlsx> <maturity_file.xlsx>")
        sys.exit(1)
    
    gap_file = sys.argv[1]
    maturity_file = sys.argv[2]
    
    print("Starting data import...")
    print("=" * 50)
    
    import_gap_assessments(gap_file)
    print()
    import_maturity_assessments(maturity_file)
    
    print("=" * 50)
    print("Import completed!")
